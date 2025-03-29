import pandas as pd
import openpyxl
from typing import Union, Optional, Tuple, List, Dict, Any


class ExcelExtractor:
    """
    Class untuk mengekstrak data dari file Excel dengan kemampuan 
    mendeteksi range data berdasarkan border dan parameter lainnya.
    """
    
    def __init__(
        self,
        file_path: str,
        sheet_name: Union[str, int] = 0,
        header_row: Optional[int] = None,
        data_start_row: Optional[int] = None,
        data_end_row: Optional[int] = None,
        header_start_col: Optional[int] = None,
        header_end_col: Optional[int] = None,
        auto_detect_range: bool = True
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.data_end_row = data_end_row
        self.header_start_col = header_start_col
        self.header_end_col = header_end_col
        self.auto_detect_range = auto_detect_range
        self.workbook = None
        self.sheet = None
        self.border_info = {}
    
    def _load_workbook(self):
        """Load workbook and sheet."""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        self.sheet = self.workbook[self.sheet_name] if isinstance(self.sheet_name, str) else self.workbook.worksheets[self.sheet_name]
    
    def _detect_range(self):
        """Detect table range based on cell content and borders."""
        if not self.sheet:
            self._load_workbook()
            
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        
        # Auto-detect header row if not specified
        if self.header_row is None and self.auto_detect_range:
            # Logic for header detection (can be enhanced)
            # For now we'll use row 0 as default
            self.header_row = 0
            
        # Auto-detect data start row if not specified
        if self.data_start_row is None and self.auto_detect_range:
            self.data_start_row = self.header_row + 1
        
        # Detect rows with bottom borders
        rows_with_bottom_border = {}
        for row in range(1, max_row + 1):
            border_cols = []
            for col in range(1, max_col + 1):
                cell = self.sheet.cell(row=row, column=col)
                if cell.border and cell.border.bottom and cell.border.bottom.style:
                    border_cols.append(col)
            if border_cols:
                rows_with_bottom_border[row] = border_cols
        
        # Find continuous ranges with bottom borders
        # Group border rows by continuous ranges of column coverage
        border_ranges = []
        for row, cols in rows_with_bottom_border.items():
            # Consider only rows with significant border coverage
            # (e.g., more than 3 columns with borders)
            if len(cols) > 3:
                border_ranges.append(row)
        
        # Find the first row after header_row with significant bottom borders
        data_borders = [r for r in border_ranges if r > (self.header_row + 1)]
        
        # Auto-detect data end row based on bottom borders
        # Use the last row with significant bottom borders
        if self.data_end_row is None and self.auto_detect_range and data_borders:
            self.data_end_row = data_borders[-1]
            
        # Save border info for debugging
        self.border_info = {
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "data_end_row": self.data_end_row,
            "max_row": max_row,
            "max_col": max_col,
            "rows_with_bottom_border": rows_with_bottom_border,
            "border_ranges": border_ranges
        }
        
    def _get_cell_value(self, row, col):
        """Get cell value safely."""
        cell = self.sheet.cell(row=row, column=col)
        return cell.value
    
    def extract(self) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Extract data from Excel file based on specified or auto-detected range.
        Returns the extracted DataFrame and border information.
        """
        # Load workbook if not already loaded
        if not self.sheet:
            self._load_workbook()
        
        # Detect range if using auto-detection
        if self.auto_detect_range:
            self._detect_range()
        
        # Use user-specified values if provided
        header_row = self.header_row if self.header_row is not None else 1
        data_start_row = self.data_start_row if self.data_start_row is not None else header_row + 1
        data_end_row = self.data_end_row if self.data_end_row is not None else self.sheet.max_row
        
        # Adjust for Excel 1-based indexing
        header_row_index = header_row
        data_start_row_index = data_start_row
        data_end_row_index = data_end_row
        
        # Determine column range
        start_col = self.header_start_col if self.header_start_col is not None else 1
        end_col = self.header_end_col if self.header_end_col is not None else self.sheet.max_column
        
        # Extract headers
        headers = []
        for col in range(start_col, end_col + 1):
            header_value = self._get_cell_value(header_row_index, col)
            headers.append(f"Column_{col}" if header_value is None or str(header_value).strip() == "" else str(header_value))
        
        # Extract data rows
        data = []
        for row in range(data_start_row_index, data_end_row_index + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(self._get_cell_value(row, col))
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        return df, self.border_info


class CSVExtractor:
    """Class untuk mengekstrak data dari file CSV."""
    
    def __init__(
        self,
        file_path: str,
        encoding: str = "utf-8",
        header_row: int = 0,
        data_start_row: int = 1,
        data_end_row: Optional[int] = None
    ):
        self.file_path = file_path
        self.encoding = encoding
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.data_end_row = data_end_row
    
    def extract(self) -> pd.DataFrame:
        """Extract data from CSV file."""
        # Read the CSV file
        df = pd.read_csv(self.file_path, encoding=self.encoding, header=None)
        
        # Extract headers
        headers = [
            f"Column_{i}" if pd.isna(h) or str(h).strip() == "" else str(h) 
            for i, h in enumerate(df.iloc[self.header_row])
        ]
        
        # Extract data rows
        data_rows = df.iloc[
            self.data_start_row:(self.data_end_row if self.data_end_row is not None else None)
        ].values
        
        # Create final DataFrame
        return pd.DataFrame(data_rows, columns=headers)


def create_extractor(
    file_path: str,
    sheet_name: Union[str, int] = 0,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    data_end_row: Optional[int] = None,
    header_start_col: Optional[int] = None,
    header_end_col: Optional[int] = None,
    auto_detect_range: bool = True,
    encoding: str = "utf-8"
) -> Union[ExcelExtractor, CSVExtractor]:
    """Factory function to create appropriate extractor based on file type."""
    file_ext = file_path.split(".")[-1].lower()
    
    if file_ext in ["xlsx", "xls", "xlsm"]:
        return ExcelExtractor(
            file_path=file_path,
            sheet_name=sheet_name,
            header_row=header_row,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
            header_start_col=header_start_col,
            header_end_col=header_end_col,
            auto_detect_range=auto_detect_range
        )
    elif file_ext == "csv":
        return CSVExtractor(
            file_path=file_path,
            encoding=encoding,
            header_row=header_row if header_row is not None else 0,
            data_start_row=data_start_row if data_start_row is not None else 1,
            data_end_row=data_end_row
        )
    else:
        raise ValueError(f"Unsupported file extension: {file_ext}")
